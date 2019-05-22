"""
    The code was developed to automate EPC simulation from gbXML file. Before you use the code, please read the tutorial file before you use.
    If you have questions about the code, Contact Yun Joon Jung (yjung77@gatech.edu) or Mengyue Liao (myliao@gatech.edu)
    # 0. Import necessary modules
    # 1. Validate xml instance with xsd
    # 2. Units check (area, volume, temperature)
    # 3. Volume
    # 4. Air change per hour (ACH)
    # 5. Thermal zone info
    # 6. Temperature schedule
    # 7. Zone schedule (Occupancy, Appliance and Lighting)
    # 8. Opaque wall area
    # 9. Below grade area
    # 10. Window area
    # 11. Roof area 
    # 12. Material properties (U-value, absoroption coefficient, solar transmittance)
    # 13. Heat capacity
    # 14. Write the results into EPC file
"""

from Tkinter import END, Label, Entry, Button, StringVar
import tkFileDialog
import tkMessageBox
import Tkinter as tk
from openpyxl import load_workbook

def _convert(xmlpath, xsdpath, epcpath):
    print "saving to", epcpath
    #######################################################################################################
    # 0. Import necessary modules
    from lxml import etree
    import numpy as np

    np.set_printoptions(linewidth=200)

    #######################################################################################################
    # 1. Validate xml instance with xsd
    xmlschema_doc = etree.parse(xsdpath)
    xmlschema = etree.XMLSchema(xmlschema_doc)
    xml_doc = etree.parse(xmlpath)

    if xmlschema.validate(xml_doc):
        print 'Valid xml'
    else:
        print 'Invalid xml'
    #######################################################################################################
    # 2. Units check (area, volume, temperature)
    url = "http://www.gbxml.org/schema"  # Define namespaces

    lenUnit = bool(xml_doc.xpath("/ns:gbXML[@lengthUnit='Feet']", namespaces={'ns': url}))
    areaUnit = bool(xml_doc.xpath("/ns:gbXML[@areaUnit='SquareFeet']", namespaces={'ns': url}))
    volUnit = bool(xml_doc.xpath("/ns:gbXML[@volumeUnit='CubicFeet']", namespaces={'ns': url}))
    temUnit = bool(xml_doc.xpath("/ns:gbXML[@temperatureUnit='F']", namespaces={'ns': url}))

    if lenUnit is True:
        lenUnit = 'IP'  # IP = Inch-pound unit
    else:
        lenUnit = 'SI'  # SI unit

    if areaUnit is True:
        areaUnit = 'IP'
    else:
        areaUnit = 'SI'

    if volUnit is True:
        volUnit = 'IP'
    else:
        volUnit = 'SI'

    if temUnit is True:
        temUnit = 'F'  # Farenheit
    else:
        temUnit = 'C'  # Celsius
    #######################################################################################################
    # 3. Volume
    numSpace = len(xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']", namespaces={'ns': url}))

    bldVolume = []  # List of each space's volume

    for i in range(0, numSpace):
        volume = float(xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/ns:Volume", namespaces={'ns': url})[i].text)
        bldVolume.append(volume)

    sumVolume = sum(bldVolume)

    if volUnit == 'IP':
        sumVolume = sumVolume * 0.0283168  # ft3 to m3
    print "\n", "Building total ventilated volume is %.1f m3" % (sumVolume)
    #######################################################################################################
    # 4. Air change per hour (ACH)
    try:
        numZone = len(xml_doc.xpath("//ns:Zone", namespaces={'ns': url}))  # Number of thermal zones
        ach = 0

        for i in range(0, numZone):
            ach += float(xml_doc.xpath("//ns:Zone/ns:AirChangesPerHour", namespaces={'ns': url})[i].text)
        ach /= numZone

        print "\n", "ACH is %0.1f (1/h). If ACH is 0.0, it is because Revit 2018 doesn't generate ACH correctly" % ach
        
    except:
        print("If you have an error, it is because Revit didn't generate ACH value(bug). Please do hand calculation")
    #######################################################################################################
    # 5. Thermal zone info
    try:
        numSpace = len(xml_doc.xpath("//ns:Space[@conditionType!='Unconditioned']", namespaces={'ns': url})) # Number of spaces
        zoneInfo = []

    # 5.1 Parse zone Id's
        for i in range(0, numSpace):
            zoneIdRef = xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@zoneIdRef", namespaces={'ns': url})[i]
            zoneInfo.append(zoneIdRef)

        zoneInfoList = list(set(zoneInfo))  # Cut the overlapped thermal zones
        zoneInfoList.sort()

        if len(zoneInfoList) > 10:  # Maximum number of thermal zones in EPC = 10
            print ("Maximum number of thermal zones in EPC is 10 (Currently, there are %d thermal zones). please reduce the number of thermal zones" % len(zoneInfo))

    # 5.2 Parse each zone's name
        numZone = len(xml_doc.xpath("//ns:Zone", namespaces={'ns': url}))  # Number of thermal zones
        zoneName = []
        for i in range(0, numZone):
            zoneName.append(xml_doc.xpath("//ns:Zone/ns:Name", namespaces={'ns': url})[i].text)

    # 5.3 Thermal zone detail information
        zone = np.zeros((7, len(zoneInfoList)), dtype=float)
        zoneRepo = []

        for i in range(0, numSpace):
            zoneRepo.append(xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@zoneIdRef", namespaces={'ns': url})[i])
            zoneRepo.append(float(xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/ns:Area", namespaces={'ns': url})[i].text))
            zoneRepo.append(float(xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/ns:PeopleHeatGain[@heatGainType='Total']",namespaces={'ns': url})[i].text))
            zoneRepo.append(float(xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/ns:EquipPowerPerArea", namespaces={'ns': url})[i].text))
            zoneRepo.append(float(xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/ns:LightPowerPerArea", namespaces={'ns': url})[i].text))
            for j in range(0, len(zoneInfoList)):
                if zoneRepo[0] == zoneInfoList[j]:
                    zone[0, j] += zoneRepo[1]  # Area
                    zone[2, j] += zoneRepo[2]  # Metabolic
                    zone[3, j] += zoneRepo[3]  # Equipment
                    zone[4, j] += zoneRepo[4]  # Light
            zoneRepo = []

        for i in range(0, len(zoneInfoList)): # Outdoor air
            zone[5, i] = float(xml_doc.xpath("//ns:Zone/ns:OAFlowPerPerson", namespaces={'ns': url})[i].text)

        if xml_doc.xpath("//ns:Zone/ns:OAFlowPerPerson/@unit", namespaces={'ns': url})[0] == 'CFM':
            for i in range(0, len(zoneInfoList)):
                zone[5, i] *= 0.47194745  # CFM/ft2/person to L/S/person

    # 5.3 Average overlapped values
        a = []
        for i in range(0, len(zoneInfoList)):
            a.append(zoneInfo.count(zoneInfoList[i]))

        for i in range(0, len(zoneInfoList)):
            if a[i] > 1:
                zone[2, i] /= a[i]
                zone[3, i] /= a[i]
                zone[4, i] /= a[i]

    # 5.5 Change units to SI unit
        if areaUnit == 'IP':
            for i in range(0, len(zoneInfoList)):
                zone[0, i] *= 0.092903
        if xml_doc.xpath("//ns:PeopleHeatGain[@heatGainType='Total']/@unit", namespaces={'ns': url})[0] == 'BtuPerHourPerson':
            for i in range(0, len(zoneInfoList)):
                zone[2, i] *= 0.293071  # BTU/hr to Watts
        if xml_doc.xpath("//ns:EquipPowerPerArea/@unit", namespaces={'ns': url})[0] == 'WattPerSquareFoot':
            for i in range(0, len(zoneInfoList)):
                zone[3, i] /= 0.092903  # ft2 to m2
        if xml_doc.xpath("//ns:LightPowerPerArea/@unit", namespaces={'ns': url})[0] == 'WattPerSquareFoot':
            for i in range(0, len(zoneInfoList)):
                zone[4, i] /= 0.092903  # ft2 to m2
        print '\n', "Zone", '\n', zoneName, '\n', zone
        
    except:
        print ("If you have errors here, please make sure you set up space and thermal zone correctly")
    #######################################################################################################
    # 6. Temperature schedule
    try:
        schTemp = np.zeros((24, 4))  # Temperature schedule
        setRepo = np.zeros((2, len(zoneInfoList)), dtype=float)

        for i in range(0, len(zoneInfoList)):
            setRepo[0, i] = float(xml_doc.xpath("//ns:Zone/ns:DesignHeatT", namespaces={'ns': url})[i].text)
            setRepo[1, i] = float(xml_doc.xpath("//ns:Zone/ns:DesignCoolT", namespaces={'ns': url})[i].text)

        setPoint = np.mean(setRepo, axis=1, dtype=float).reshape(2, 1)

        if xml_doc.xpath("//ns:Zone/ns:DesignHeatT/@unit", namespaces={'ns': url})[0] == 'F':
            setPoint[0, 0] = (setPoint[0, 0] - 32) * 0.555555  # F to C
            setPoint[1, 0] = (setPoint[1, 0] - 32) * 0.555555

        for i in range(0, 24):  # Lcoate set temperatures
            if i <= 7:
                schTemp[i, 0] = 20
                schTemp[i, 1] = 20
                schTemp[i, 2] = 25
                schTemp[i, 3] = 25
            elif i > 7 and i <= 17:
                schTemp[i, 0] = setPoint[0, 0]
                schTemp[i, 1] = 20
                schTemp[i, 2] = setPoint[1, 0]
                schTemp[i, 3] = 25
            elif i > 17:
                schTemp[i, 0] = 20
                schTemp[i, 1] = 20
                schTemp[i, 2] = 25
                schTemp[i, 3] = 25

        print '\n' + "Temperature set-point schedule (Unit: deg C)", '\n', schTemp
        
    except:
        print("If you have errors here, please make sure you set up heating/cooling set-point temperature")
    #######################################################################################################
    # 7. Zone schedule (Occupancy, Appliance and Lighting)
    try:
        numSch = len(xml_doc.xpath("//ns:Schedule", namespaces={'ns': url}))  # Number of schedule
        numWeekSch = len(xml_doc.xpath("//ns:WeekSchedule", namespaces={'ns': url}))  # Number of weekScheudle
        numDaySch = len(xml_doc.xpath("//ns:DaySchedule", namespaces={'ns': url}))  # Number of dayScheudle

        equipRepo = [] # Equipment Schedule repository
        for i in range(0, numSpace):
            equipRepo.append(xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@peopleScheduleIdRef", namespaces={'ns': url})[i])

        equipRepo = list(set(equipRepo))
        equipRepo.sort()

        equipInfo = [] 
        for i in range(0, numSch):
            for j in range(0, len(equipRepo)):
                if xml_doc.xpath("//ns:Schedule/@id", namespaces={'ns': url})[i] == equipRepo[j]:
                    equipInfo.append(xml_doc.xpath("//ns:Schedule/ns:YearSchedule/ns:WeekScheduleId/@weekScheduleIdRef", namespaces={'ns': url})[i])

        equipSet = []
        for i in range(0, numWeekSch):
            for j in range(0, len(equipInfo)):
                if xml_doc.xpath("//ns:WeekSchedule/@id", namespaces={'ns': url})[i] == equipInfo[j]:
                    equipSet.append(xml_doc.xpath("//ns:WeekSchedule/ns:Day/@dayScheduleIdRef", namespaces={'ns': url})[i])

        equipSch = np.zeros((24, len(equipSet) * 2), dtype=float) # equip schedule table
        for i in range(0, numDaySch):
            for j in range(0, len(equipSet)):
                if xml_doc.xpath("//ns:DaySchedule/@id", namespaces={'ns': url})[i] == equipSet[j]:
                    for k in range(0, 24):
                        a = equipSet[j]
                        if i == 0:
                            equipSch[k, j] = float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)
                            equipSch[k, j + 1] = 0.2 * float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)
                        elif i > 0:
                            equipSch[k, j * 2] = float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)
                            equipSch[k, j * 2 + 1] = 0.2 * float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)

    # Appliance schedule
        appRepo = []

        for i in range(0, numSpace):
            appRepo.append(xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@equipmentScheduleIdRef", namespaces={'ns': url})[i])

        appRepo = list(set(appRepo))
        appRepo.sort()

        appInfo = []
        for i in range(0, numSch):
            for j in range(0, len(appRepo)):
                if xml_doc.xpath("//ns:Schedule/@id", namespaces={'ns': url})[i] == appRepo[j]:
                    appInfo.append(xml_doc.xpath("//ns:Schedule/ns:YearSchedule/ns:WeekScheduleId/@weekScheduleIdRef",namespaces={'ns': url})[i])

        appSet = []
        for i in range(0, numWeekSch):
            for j in range(0, len(equipInfo)):
                if xml_doc.xpath("//ns:WeekSchedule/@id", namespaces={'ns': url})[i] == appInfo[j]:
                    appSet.append(xml_doc.xpath("//ns:WeekSchedule/ns:Day/@dayScheduleIdRef", namespaces={'ns': url})[i])

        appSch = np.zeros((24, len(appSet) * 2), dtype=float)  # equip schedule table
        for i in range(0, numDaySch):
            for j in range(0, len(appSet)):
                if xml_doc.xpath("//ns:DaySchedule/@id", namespaces={'ns': url})[i] == appSet[j]:
                    for k in range(0, 24):
                        a = appSet[j]
                        if i == 0:
                            appSch[k, j] = float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)
                            appSch[k, j + 1] = 0.2 * float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)
                        elif i > 0:
                            appSch[k, j * 2] = float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)
                            appSch[k, j * 2 + 1] = 0.2 * float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)

    # Lighting schedule
        ligRepo = []

        for i in range(0, numSpace):
            ligRepo.append(xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@lightScheduleIdRef", namespaces={'ns': url})[i])

        ligRepo = list(set(ligRepo))
        ligRepo.sort()

        ligInfo = []
        for i in range(0, numSch):
            for j in range(0, len(ligRepo)):
                if xml_doc.xpath("//ns:Schedule/@id", namespaces={'ns': url})[i] == ligRepo[j]:
                    ligInfo.append(xml_doc.xpath("//ns:Schedule/ns:YearSchedule/ns:WeekScheduleId/@weekScheduleIdRef",namespaces={'ns': url})[i])

        ligSet = []
        for i in range(0, numWeekSch):
            for j in range(0, len(equipInfo)):
                if xml_doc.xpath("//ns:WeekSchedule/@id", namespaces={'ns': url})[i] == ligInfo[j]:
                    ligSet.append(xml_doc.xpath("//ns:WeekSchedule/ns:Day/@dayScheduleIdRef", namespaces={'ns': url})[i])

        ligSch = np.zeros((24, len(ligSet) * 2), dtype=float)  # light schedule table
        for i in range(0, numDaySch):
            for j in range(0, len(equipSet)):
                if xml_doc.xpath("//ns:DaySchedule/@id", namespaces={'ns': url})[i] == ligSet[j]:
                    for k in range(0, 24):
                        a = ligSet[j]
                        if i == 0:
                            ligSch[k, j] = float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)
                            ligSch[k, j + 1] = 0.2 * float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)
                        elif i > 0:
                            ligSch[k, j * 2] = float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)
                            ligSch[k, j * 2 + 1] = 0.2 * float(xml_doc.xpath("//ns:DaySchedule[@id='%s']/ns:ScheduleValue" % (a), namespaces={'ns': url})[k].text)

    # Distribtue to total zone level
        schZone = np.zeros((24, len(zoneInfoList) * 6), dtype=float)  # Zone schedule

        for i in range(0, numSpace):
            for j in range(0, len(zoneInfoList)):
                if xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@zoneIdRef", namespaces={'ns': url})[i] == zoneInfoList[j]:
                    for k in range(0, len(equipRepo)):
                        if xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@peopleScheduleIdRef",namespaces={'ns': url})[j] == equipRepo[k]:
                            if j == 0:
                                schZone[:, j] += equipSch[:, k]
                                schZone[:, j + 1] += equipSch[:, k + 1]
                            if j > 0:
                                schZone[:, j * 6] += equipSch[:, k]
                                schZone[:, j * 6 + 1] += equipSch[:, k + 1]

        for i in range(0, numSpace):
            for j in range(0, len(zoneInfoList)):
                if xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@zoneIdRef", namespaces={'ns': url})[i] == zoneInfoList[j]:
                    for k in range(0, len(appRepo)):
                        if xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@equipmentScheduleIdRef",namespaces={'ns': url})[j] == appRepo[k]:
                            if j == 0:
                                schZone[:, j + 2] += appSch[:, k]
                                schZone[:, j + 3] += appSch[:, k + 1]
                            if j > 0:
                                schZone[:, (j * 6) + 2] += appSch[:, k]
                                schZone[:, (j * 6) + 3] += appSch[:, k + 1]

        for i in range(0, numSpace):
            for j in range(0, len(zoneInfoList)):
                if xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@zoneIdRef", namespaces={'ns': url})[i] == zoneInfoList[j]:
                    for k in range(0, len(ligRepo)):
                        if xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@lightScheduleIdRef", namespaces={'ns': url})[j] == ligRepo[k]:
                            if j == 0:
                                schZone[:, j + 4] += ligSch[:, k]
                                schZone[:, j + 5] += ligSch[:, k + 1]
                            if j > 0:
                                schZone[:, (j * 6) + 4] += ligSch[:, k]
                                schZone[:, (j * 6) + 5] += ligSch[:, k + 1]

    # Divide by the number of overlapped schedules
        schOverlap = np.zeros((1, len(zoneInfoList)), dtype=int)

        for i in range(0, numSpace):
            for j in range(0, len(zoneInfoList)):
                if xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Building/ns:Space[@conditionType!='Unconditioned']/@zoneIdRef", namespaces={'ns': url})[i] == zoneInfoList[j]:
                    schOverlap[0, j] += 1
        for i in range(0, len(zoneInfoList)):
            schZone[:, 6 * i:6 * i + 5] /= schOverlap[0, i]

        print '\n' + "Zone schedule", '\n', schZone
        
    except:
        print("If you have errors here, please make sure you set up schedule information correctly")
    #######################################################################################################
    # 8. Opaque wall area
    try:
        numWall = len(xml_doc.xpath("//ns:Surface[@constructionIdRef and @surfaceType='ExteriorWall']", namespaces={'ns': url}))  # Number of opaque wall surfaces
        wallInfo = []  # Construction information repository
        wallIdSet = []  # Collect construction Ids and areas
        ex_wall = {}

        for i in range(0, numWall):
            wallIdRef = xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Surface[@constructionIdRef and @surfaceType='ExteriorWall']/@constructionIdRef",namespaces={'ns': url})[i]
            wallInfo.append(wallIdRef)
            azimuth = float(xml_doc.xpath("//ns:Surface[@constructionIdRef and @surfaceType='ExteriorWall']//ns:Azimuth",namespaces={'ns': url})[i].text)
            width = float(xml_doc.xpath("//ns:Surface[@constructionIdRef and @surfaceType='ExteriorWall']//ns:Width",namespaces={'ns': url})[i].text)
            height = float(xml_doc.xpath("//ns:Surface[@constructionIdRef and @surfaceType='ExteriorWall']//ns:Height",namespaces={'ns': url})[i].text)
            area = width * height
            wallIdSet.append([wallIdRef, area])
            ex_wall['Wall %s' % (i + 1)] = [azimuth, area]

        north = []  # North elevation area list
        northEast = []
        east = []
        northWest = []
        south = []
        southEast = []
        southWest = []
        west = []

        for wall in range(0, numWall):
            n = ex_wall['Wall %s' % (wall + 1)]
            if n[0] >= -22.5 + 360 or n[0] < 22.5:  # North Orientation
                north.append(n[1])
            elif n[0] >= 22.5 and n[0] < 67.5:
                northEast.append(n[1])
            elif n[0] >= 67.5 and n[0] < 67.5 + 45:
                east.append(n[1])
            elif n[0] >= 67.5 + 45 and n[0] < 67.5 + 90:
                southEast.append(n[1])
            elif n[0] >= 67.5 + 90 and n[0] < 67.5 + 135:
                south.append(n[1])
            elif n[0] >= 67.5 + 135 and n[0] < 67.5 + 180:
                southWest.append(n[1])
            elif n[0] >= 67.5 + 180 and n[0] < 67.5 + 180 + 45:
                west.append(n[1])

        northWall = sum(north)  # Aggregate entire north oriented-wall areas
        northEastWall = sum(northEast)
        eastWall = sum(east)
        northWestWall = sum(northWest)
        southWall = sum(south)
        southEastWall = sum(southEast)
        southWestWall = sum(southWest)
        westWall = sum(west)

        if areaUnit == 'IP':
            northWall = northWall * 0.092903
            northEastWall = northEastWall * 0.092903
            eastWall = eastWall * 0.092903
            northWestWall = northWestWall * 0.092903
            southWall = southWall * 0.092903
            southEastWall = southEastWall * 0.092903
            southWestWall = southWestWall * 0.092903
            westWall = westWall * 0.092903

        print '\n', "***** Opauqe wall area *****"
        print "North: %.1f m2" % northWall
        print "Northeast: %.1f m2" % northEastWall
        print "East wall: %.1f m2" % eastWall
        print "NorthWest: %.1f m2" % northWestWall
        print "South: %.1f m2" % southWall
        print "SouthEast: %.1f m2" % southEastWall
        print "SouthWest: %.1f m2" % southWestWall
        print "West: %.1f m2" % westWall
        print "****************************"
        
    except:
        print("If you have errors here, please make sure your model geometry was correctly built")
    #######################################################################################################
    # 9. Below grade area
    try:
        numBel = len(xml_doc.xpath("//ns:Surface[@surfaceType='UndergroundWall']", namespaces={'ns': url}))  # Number of Below grade surfaces
        belGradeArea = 0
        area = 0
        for i in range(0, numBel):
            width = float(xml_doc.xpath("//ns:Surface[@surfaceType='UndergroundWall']//ns:Width",namespaces={'ns': url})[i].text)
            height = float(xml_doc.xpath("//ns:Surface[@surfaceType='UndergroundWall']//ns:Height",namespaces={'ns': url})[i].text)
            area = width * height
            belGradeArea = belGradeArea + area
            
        if areaUnit == 'IP':
            belGradeArea = belGradeArea * 0.092903
            
        print '\n', "***** Below grade area *****"
        print "Below grade: %.1f m2" % belGradeArea
        print "****************************"

    except:
        print("Please check underground part in Revit")
    #######################################################################################################
    # 10. Window area
    try:
        numWin = len(xml_doc.xpath("//ns:Opening[@openingType!='NonSlidingDoor' and @openingType!='SlidingDoor' and @openingType!='FixedSkylight' and @openingType!='FixedSkylight' and @openingType!='Air']",namespaces={'ns': url}))
        winInfo = []  # Construction information repository
        winIdSet = []  # Collect construction Ids and areas
        winSet = {}

        for i in range(0, numWin):
            winIdRef = xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Surface[@surfaceType='ExteriorWall']/ns:Opening/@windowTypeIdRef",namespaces={'ns': url})[i]
            winInfo.append(winIdRef)

            azimuth = float(xml_doc.xpath("//ns:Opening[@windowTypeIdRef and @openingType!='NonSlidingDoor' and @openingType!='SlidingDoor' and @openingType!='FixedSkylight' and @openingType!='FixedSkylight' and @openingType!='Air']/ns:RectangularGeometry/ns:Azimuth", namespaces={'ns': url})[i].text)
            width = float(xml_doc.xpath("//ns:Opening[@windowTypeIdRef and @openingType!='NonSlidingDoor' and @openingType!='SlidingDoor' and @openingType!='FixedSkylight' and @openingType!='FixedSkylight' and @openingType!='Air']/ns:RectangularGeometry/ns:Width", namespaces={'ns': url})[i].text)
            height = float(xml_doc.xpath("//ns:Opening[@windowTypeIdRef and @openingType!='NonSlidingDoor' and @openingType!='SlidingDoor' and @openingType!='FixedSkylight' and @openingType!='FixedSkylight' and @openingType!='Air']/ns:RectangularGeometry/ns:Height", namespaces={'ns': url})[i].text)
            area = width * height

            winIdSet.append([winIdRef, area])
            winSet['Window %s' % (i + 1)] = [azimuth, area]

        northWin = []  # north window area lists
        northEastWin = []
        eastWin = []
        northWestWin = []
        southWin = []
        southEastWin = []
        southWestWin = []
        westWin = []

        for win in range(0, numWin):
            n = winSet['Window %s' % (win + 1)]
            if n[0] >= -22.5 + 360 or n[0] < 22.5:
                northWin.append(n[1])
            elif n[0] >= 22.5 and n[0] < 67.5:
                northEastWin.append(n[1])
            elif n[0] >= 67.5 and n[0] < 67.5 + 45:
                eastWin.append(n[1])
            elif n[0] >= 67.5 + 45 and n[0] < 67.5 + 90:
                southEastWin.append(n[1])
            elif n[0] >= 67.5 + 90 and n[0] < 67.5 + 135:
                southWin.append(n[1])
            elif n[0] >= 67.5 + 135 and n[0] < 67.5 + 180:
                southWestWin.append(n[1])
            elif n[0] >= 67.5 + 180 and n[0] < 67.5 + 180 + 45:
                westWin.append(n[1])

        northWin = sum(northWin)
        northEastWin = sum(northEastWin)
        eastWin = sum(eastWin)
        northWestWin = sum(northWestWin)
        southWin = sum(southWin)
        southEastWin = sum(southEastWin)
        southWestWin = sum(southWestWin)
        westWin = sum(westWin)

        #Skylight
        numSky = len(xml_doc.xpath("//ns:Opening[@openingType='FixedSkylight' or @openingType='OperableSkylight']",namespaces={'ns': url}))

        skyArea = 0
        area = 0
        for i in range(0,numSky):
            width = float(xml_doc.xpath("//ns:Opening[@windowTypeIdRef and @openingType='FixedSkylight' or @openingType='OperableSkylight']/ns:RectangularGeometry/ns:Width", namespaces={'ns': url})[i].text)
            height = float(xml_doc.xpath("//ns:Opening[@windowTypeIdRef and @openingType='FixedSkylight' or @openingType='OperableSkylight']/ns:RectangularGeometry/ns:Height", namespaces={'ns': url})[i].text)
            area = width * height
            skyArea = skyArea + area
    
        if areaUnit == 'IP':
            northWin = northWin * 0.092903
            northEastWin = northEastWin * 0.092903
            eastWin = eastWin * 0.092903
            northWestWin = northWestWin * 0.092903
            southWin = southWin * 0.092903
            southEastWin = southEastWin * 0.092903
            southWestWin = southWestWin * 0.092903
            westWin = westWin * 0.092903
            skyArea = skyArea * 0.092903

        print '\n' + "******* Window area *******"
        print "North: %.1f m2" % northWin
        print "Northeast: %.1f m2" % northEastWin
        print "East: %.1f m2" % eastWin
        print "NorthWest: %.1f m2" % northWestWin
        print "SouthWall: %.1f m2" % southWin
        print "SouthEast: %.1f m2" % southEastWin
        print "SouthWest: %.1f m2" % southWestWin
        print "West: %.1f m2" % westWin
        print "Skylight: %.1f m2" % skyArea
        print "****************************"
        
    except:
        msg = "Please make sure your model window geometry was correctly built. Some Revit windows(e.g. round-shape window and curtain wall door) have bugs. Please contact the author"
        print(msg)
    #######################################################################################################
    # 11. Roof area
    try:
        numRoof = len(xml_doc.xpath("//ns:Surface[@surfaceType='Roof']", namespaces={'ns': url}))
        roofInfo = []  # Construction information repository
        roofIdSet = []  # Coolect construction Ids and areas

        setRoofArea = []

        for i in range(0, numRoof):
            roofIdRef = xml_doc.xpath("/ns:gbXML/ns:Campus/ns:Surface[@surfaceType='Roof']/@constructionIdRef", namespaces={'ns': url})[i]
            roofInfo.append(roofIdRef)

            azimuth = float(xml_doc.xpath("//ns:Surface[@surfaceType='Roof']/ns:RectangularGeometry/ns:Azimuth", namespaces={'ns': url})[i].text)
            width = float(xml_doc.xpath("//ns:Surface[@surfaceType='Roof']/ns:RectangularGeometry/ns:Width", namespaces={'ns': url})[i].text)
            height = float(xml_doc.xpath("//ns:Surface[@surfaceType='Roof']/ns:RectangularGeometry/ns:Height", namespaces={'ns': url})[i].text)
            roofArea = width * height
            setRoofArea.append(roofArea)

            roofIdSet.append([roofIdRef, roofArea])

        roofArea = sum(setRoofArea)

        if areaUnit == 'IP':
            roofArea = roofArea * 0.092903

        print '\n' + "******** Roof area ********"
        print "Roof: %.1f m2" % roofArea
        print "***************************"
    except:
        print("If you have errors here, please make sure your model geometry was correctly built")
    #######################################################################################################
    # 12. Material properties (U-value, absoroption coefficient, solar transmittance)
    try:
    # 12.1 Roof
        roofSort = (list(set(roofInfo)))
        roof_Uvalue = {}  # Roof U-values' repository
        roof_Absorp = {}

        for i in range(0, len(roofSort)):
            Uvalue = float(xml_doc.xpath("//ns:Construction[@id = '%s']/ns:U-value" % str(roofSort[i]), namespaces={'ns': url})[0].text)
            absorp = float(xml_doc.xpath("//ns:Construction[@id = '%s']/ns:Absorptance" % str(roofSort[i]), namespaces={'ns': url})[0].text)  # Absorptance
            if bool(str(xml_doc.xpath("//ns:Construction[@id = '%s']/ns:U-value[@unit ='WPerSquareMeterK']" % str(roofSort[i]), namespaces={'ns': url}))) == False:
                Uvalue = Uvalue * 5.678  # Covert to SI Units
            roof_Uvalue['%s' % str(roofSort[i])] = Uvalue
            roof_Absorp['%s' % str(roofSort[i])] = absorp

    # 12.2 Calculate overall Uvalue ((U1A1+U2A2+...)/(A1+A2+...)) and absorptance
        UA = 0
        for i in range(0, len(roofIdSet)):
            for j in range(0, len(roofSort)):
                if roofIdSet[i][0] == roofSort[j]:
                    if areaUnit == 'IP':
                        UA += roof_Uvalue.get(roofSort[j]) * roofIdSet[i][1] * 0.092903  # U value * area
                    else:
                        UA += roof_Uvalue.get(roofSort[j]) * roofIdSet[i][1]
        roofUvalue = UA / roofArea

        aA = 0  # Absorptance * Area
        for i in range(0, len(roofIdSet)):
            for j in range(0, len(roofSort)):
                if roofIdSet[i][0] == roofSort[j]:
                    if areaUnit == 'IP':
                        aA += roof_Absorp.get(roofSort[j]) * roofIdSet[i][1] * 0.092903
                    else:
                        aA += roof_Absorp.get(roofSort[j]) * roofIdSet[i][1]
        roofAbsor = aA / roofArea 

        print '\n' + "***** Construction Info *******"
        print "Roof U-Value: %0.2f W/(m2K)" % roofUvalue
        print "Roof absorption coefficient: %0.2f" % roofAbsor

    # 12.3 Wall 
        wallSort = (list(set(wallInfo)))
        wall_Uvalue = {}  # Roof U-values' repository
        wall_Absorp = {}

        for i in range(0, len(wallSort)):
            Uvalue = float(xml_doc.xpath("//ns:Construction[@id = '%s']/ns:U-value" % str(wallSort[i]), namespaces={'ns': url})[0].text)
            absorp = float(xml_doc.xpath("//ns:Construction[@id = '%s']/ns:Absorptance" % str(wallSort[i]), namespaces={'ns': url})[0].text)  # Absorptance
            if bool(str(xml_doc.xpath("//ns:Construction[@id = '%s']/ns:U-value[@unit ='WPerSquareMeterK']" % str(wallSort[i]),namespaces={'ns': url}))) == False:
                Uvalue = Uvalue * 5.678  # Covert to SI Units
            wall_Uvalue['%s' % str(wallSort[i])] = Uvalue
            wall_Absorp['%s' % str(wallSort[i])] = absorp

    # 12.4 Calculate overall Uvalue ((U1A1+U2A2+...)/(A1+A2+...)) and absorptance
        UA = 0
        for i in range(0, len(wallIdSet)):
            for j in range(0, len(wallSort)):
                if wallIdSet[i][0] == wallSort[j]:
                    if areaUnit == 'IP':
                        UA += wall_Uvalue.get(wallSort[j]) * wallIdSet[i][1] * 0.092903  # U value * area
                    else:
                        UA += wall_Uvalue.get(wallSort[j]) * wallIdSet[i][1]

        wallTotArea = northWall + northEastWall + eastWall + northWestWall + southWall + southEastWall + southWestWall + westWall
        wallUvalue = UA / wallTotArea

        aA = 0  # Absorptance * Area
        for i in range(0, len(wallIdSet)):
            for j in range(0, len(wallSort)):
                if wallIdSet[i][0] == wallSort[j]:
                    if areaUnit == 'IP':
                        aA += wall_Absorp.get(wallSort[j]) * wallIdSet[i][1] * 0.092903
                    else:
                        aA += wall_Absorp.get(wallSort[j]) * wallIdSet[i][1]
        wallAbsor = aA / wallTotArea

        print "Wall U-Value: %0.2f W/(m2K)" % wallUvalue
        print "Wall absorption coefficient: %0.2f" % wallAbsor

    # 12.5 Window
        winSort = (list(set(winInfo)))
        win_Uvalue = {}  # Roof U-values' repository
        win_Absorp = {}

        for i in range(0, len(winSort)):
            Uvalue = float(xml_doc.xpath("//ns:WindowType[@id = '%s']/ns:U-value" % str(winSort[i]), namespaces={'ns': url})[0].text)
            SHGC = float(xml_doc.xpath("//ns:WindowType[@id = '%s']/ns:SolarHeatGainCoeff" % str(winSort[i]), namespaces={'ns': url})[0].text)  # Solar heat gain coefficient
            if bool(str(xml_doc.xpath("//ns:Construction[@id = '%s']/ns:U-value[@unit ='WPerSquareMeterK']" % str(winSort[i]),namespaces={'ns': url}))) == False:
                Uvalue = Uvalue * 5.678  # Covert to SI Units
            win_Uvalue['%s' % str(winSort[i])] = Uvalue
            win_Absorp['%s' % str(winSort[i])] = SHGC

    # 12.6 Calculate overall Uvalue ((U1A1+U2A2+...)/(A1+A2+...)) and absorptance
        UA = 0
        for i in range(0, len(winIdSet)):
            for j in range(0, len(winSort)):
                if winIdSet[i][0] == winSort[j]:
                    if areaUnit == 'IP':
                        UA += win_Uvalue.get(winSort[j]) * winIdSet[i][1] * 0.092903  # U value * area
                    else:
                        UA += win_Uvalue.get(winSort[j]) * winIdSet[i][1]

        winTotArea = northWin + northEastWin + eastWin + northWestWin + southWin + southEastWin + southWestWin + westWin
        winUvalue = UA / winTotArea

        SHGCA = 0  # SHGC * Area
        for i in range(0, len(winIdSet)):
            for j in range(0, len(winSort)):
                if winIdSet[i][0] == winSort[j]:
                    if areaUnit == 'IP':
                        SHGCA += win_Absorp.get(winSort[j]) * winIdSet[i][1] * 0.092903
                    else:
                        SHGCA += win_Absorp.get(winSort[j]) * winIdSet[i][1]

        winSHGC = SHGCA / winTotArea
        print "Window U-Value: %0.2f W/(m2K)" % winUvalue
        print "Window SHGC: %0.2f" % winSHGC
        print "*******************************"
    except:
       print("If you have errors here, please make sure you set up material thermal properties")
    #######################################################################################################
    # 13. Heat capacity
    try:
    # 13.1 Roof
        numCon = len(xml_doc.xpath("//ns:Construction", namespaces={'ns': url}))  # Number of constrution types
        numLay = len(xml_doc.xpath("//ns:Layer", namespaces={'ns': url}))  # Number of layers
        numMat = len(xml_doc.xpath("//ns:Material", namespaces={'ns': url}))  # Number of materials

        roofInfoList = list(set(roofInfo))
        roofRepo = []
        roofMat = []  # Roof material Id

        for i in range(0, len(roofInfoList)):  # Parse Layer Id
            for j in range(0, numCon):
                if xml_doc.xpath("//ns:Construction/@id", namespaces={'ns': url})[j] == roofInfoList[i]:
                    roofRepo.append(xml_doc.xpath("//ns:Construction/ns:LayerId/@layerIdRef", namespaces={'ns': url})[j])

        for i in range(0, len(roofRepo)):  # Parse material Ids
            for j in range(0, numLay):
                if xml_doc.xpath("//ns:Layer/@id", namespaces={'ns': url})[j] == roofRepo[i]:
                    for k in range(0, int(xml_doc.xpath("count(//ns:Layer[@id='%s']/ns:MaterialId)" % str(roofRepo[i]),namespaces={'ns': url}))):
                        roofMat.append(xml_doc.xpath("//ns:Layer[@id='%s']/ns:MaterialId/@materialIdRef" % str(roofRepo[i]),namespaces={'ns': url})[k])

        roofThermalMass = np.zeros((1, len(roofMat)), dtype=float)  # Array for storing each material heat capacity

        for i in range(0, len(roofMat)):  # Calculate Heat Capacity (=Density*Specific heat*thickness)
            density = float(xml_doc.xpath("//ns:Material[@id='%s']/ns:Density" % str(roofMat[i]), namespaces={'ns': url})[0].text)
            Cp = float(xml_doc.xpath("//ns:Material[@id='%s']/ns:SpecificHeat" % str(roofMat[i]), namespaces={'ns': url})[0].text)
            thick = float(xml_doc.xpath("//ns:Material[@id='%s']/ns:Thickness" % str(roofMat[i]), namespaces={'ns': url})[0].text)
            roofThermalMass[0, i] = density * Cp * thick

        roofHC = roofThermalMass.sum()  # Roof heat capacity

    # 13.2 Wall
        wallInfoList = list(set(wallInfo))
        wallRepo = []
        wallMat = []  # Roof material Id

        for i in range(0, len(wallInfoList)):  # Parse Layer Id
            for j in range(0, numCon):
                if xml_doc.xpath("//ns:Construction/@id", namespaces={'ns': url})[j] == wallInfoList[i]:
                    wallRepo.append(xml_doc.xpath("//ns:Construction/ns:LayerId/@layerIdRef", namespaces={'ns': url})[j])

        for i in range(0, len(wallRepo)):  # Parse material Ids
            for j in range(0, numLay):
                if xml_doc.xpath("//ns:Layer/@id", namespaces={'ns': url})[j] == wallRepo[i]:
                    for k in range(0, int(xml_doc.xpath("count(//ns:Layer[@id='%s']/ns:MaterialId)" % str(wallRepo[i]),namespaces={'ns': url}))):
                        wallMat.append(xml_doc.xpath("//ns:Layer[@id='%s']/ns:MaterialId/@materialIdRef" % str(wallRepo[i]),namespaces={'ns': url})[k])

        wallThermalMass = np.zeros((1, len(wallMat)), dtype=float)  # Array for storing each material heat capacity

        for i in range(0, len(wallMat)):  # Calculate Heat Capacity (=Density*Specific heat*thickness)
            density = float(xml_doc.xpath("//ns:Material[@id='%s']/ns:Density" % str(wallMat[i]), namespaces={'ns': url})[0].text)
            Cp = float(xml_doc.xpath("//ns:Material[@id='%s']/ns:SpecificHeat" % str(wallMat[i]), namespaces={'ns': url})[0].text)
            thick = float(xml_doc.xpath("//ns:Material[@id='%s']/ns:Thickness" % str(wallMat[i]), namespaces={'ns': url})[0].text)
            wallThermalMass[0, i] = density * Cp * thick

        wallHC = wallThermalMass.sum()  # Roof heat capacity
        HCTotal = wallHC + roofHC

        print '\n' + "Envelope heat capacity is"
        if HCTotal >= 0 and HCTotal < 95000:
            print "Very Light: 80,000*Af"
        elif HCTotal >= 95000 and HCTotal < 137500:
            print "Light: 110,000*Af"
        elif HCTotal >= 137500 and HCTotal < 212500:
            print "Medium: 165,000*Af"
        elif HCTotal >= 212500 and HCTotal < 315000:
            print "Heavy: 260,000*Af"
        elif HCTotal >= 315000:
            print "Very heavy: 370,000*Af"
            
    except:
        print("If you have errors here,  please make sure you set up material thermal properties")
    #######################################################################################################
    # 14. Write the results into EPC file
    try:
        wb = load_workbook('EPC_Input.xlsx')
        sheet = wb['INPUT']
    except:
        print("Please chech whether you have EPC_Input.xlsx file in the folder")

    # Fit all values into EPC sheet
    # Building Ventilated Volume
    sheet['C14'] = sumVolume
    
    # Heat Capacity
    try:
        HCRepository = ["Very Light: 80,000 * Af", "Light : 110,000 * Af", "Medium: 165,000 * Af", "Heavy: 260,000 * Af", "Very heavy: 370,000 * Af"  ]
        if HCTotal >= 0 and HCTotal < 95000:
            Sheet['C18'] = HCRepository[0]
        elif HCTotal >= 95000 and HCTotal < 137500:
            Sheet['C18'] = HCRepository[1]
        elif HCTotal >= 137500 and HCTotal < 212500:
            Sheet['C18'] = HCRepository[2]
        elif HCTotal >= 212500 and HCTotal < 315000:
            Sheet['C18'] = HCRepository[3]
        elif HCTotal >= 315000:
            Sheet['C18'] = HCRepository[4]
    except:
        print("Please check your heat capacity properties")
        
    # Zone
    try:
        for i in range(0,len(zoneName)): # Locate zone names and internal gains
            a = sheet.cell(row=9, column=7+i, value=zoneName[i]) # Zone names
            b = sheet.cell(row=10, column=7+i, value=zone[0,i]) # Area
            c = sheet.cell(row=12, column=7+i, value=zone[2,i]) # Metabolic rate
            d = sheet.cell(row=13, column=7+i, value=zone[3,i]) # Appliance
            e = sheet.cell(row=14, column=7+i, value=zone[4,i]) # Lighting
            f = sheet.cell(row=15, column=7+i, value=zone[5,i]) # Outdoor Air
    except:
        print("Please check whether your zone setting was correctly done")
        
    # Temp set-point schedule
    try:
        for i in range(0,24):
            for j in range(0,4):
                g = sheet.cell(row=20+i, column=7+j, value=schTemp[i,j]) 
        
    # Zone schedule (Occ,App,Light)
        for j in range(0,len(zoneName)*6):
            for i in range(0,24):
                if (j >= 0 and j < 6):
                    h = sheet.cell(row=20+i, column=13+j, value=schZone[i,j]) # Occ_WD          
                elif (j >= 6 and j < 12):
                    h = sheet.cell(row=20+i, column=15+j, value=schZone[i,j])
                elif (j >= 12 and j < 18):
                    h = sheet.cell(row=20+i, column=17+j, value=schZone[i,j])
                elif (j >= 18 and j < 24):
                    h = sheet.cell(row=20+i, column=19+j, value=schZone[i,j])
                elif (j >= 24 and j < 30):
                    h = sheet.cell(row=20+i, column=21+j, value=schZone[i,j])
                elif (j >= 30 and j < 36):
                    h = sheet.cell(row=20+i, column=23+j, value=schZone[i,j])
                elif (j >= 42 and j < 48):
                    h = sheet.cell(row=20+i, column=25+j, value=schZone[i,j])
                elif (j >= 48 and j < 54):
                    h = sheet.cell(row=20+i, column=27+j, value=schZone[i,j])
                elif (j >= 54 and j < 60):
                    h = sheet.cell(row=20+i, column=29+j, value=schZone[i,j])
                elif (j >= 60 and j < 66):
                    h = sheet.cell(row=20+i, column=31+j, value=schZone[i,j])
    except:
        print("Please check whether your zone schedule settings were correctly done")
        
    # Envelope (Opaque wall, Roof and window)
    try:
        sheet['G50'] = southWall # Opaque wall
        sheet['G51'] = southEastWall
        sheet['G52'] = eastWall
        sheet['G53'] = northEastWall
        sheet['G54'] = northWall
        sheet['G55'] = northWestWall
        sheet['G56'] = westWall
        sheet['G57'] = southWestWall
        sheet['G58'] = roofArea # Roof
        sheet['G59'] = belGradeArea # Below grade
        sheet['I50'] = southWin # Window
        sheet['I51'] = southEastWin
        sheet['I52'] = eastWin
        sheet['I53'] = northEastWin
        sheet['I54'] = northWin
        sheet['I55'] = northWestWin
        sheet['I56'] = westWin
        sheet['I57'] = southWestWin
        sheet['I58'] = skyArea
        
    except:
        print("Please check your geometry")     

    # Material
    try:
        sheet['G66'] = wallUvalue
        sheet['H66'] = wallAbsor
        sheet['G64'] = roofUvalue
        sheet['H64'] = roofAbsor
        sheet['G68'] = winUvalue
        sheet['J68'] = winSHGC
    except:
        print("Please check your material thermal properties")
            
    wb.save('EPC_Input.xlsx')

    # End of _convert function
    #######################################################################################################
def main():
    root = tk.Tk()
    root.title("gbXML to EPC converter")

    Label(root, text="gbXML file").grid(row=0)
    Label(root, text="xsd file").grid(row=1)
    Label(root, text="EPC spreadsheet file").grid(row=2)
    status = StringVar()
    Label(root, textvariable=status).grid(row=3)
    status.set("Status: Ready")

    xmlentry = Entry(root)
    xsdentry = Entry(root)
    epcentry = Entry(root)
    xmlentry.grid(row=0, column=1)
    xsdentry.grid(row=1, column=1)
    epcentry.grid(row=2, column=1)

    xmlpath = ""
    xsdpath = ""
    epcpath = ""

    def openxml():
        ftypes = [('gbXML files', '*.xml')]
        dlg = tkFileDialog.Open(filetypes=ftypes)
        fl = dlg.show()

        if fl != '':
            xmlentry.delete(0, END)
            xmlentry.insert(0, fl)

    def openxsd():
        ftypes = [('xsd files', '*.xsd')]
        dlg = tkFileDialog.Open(filetypes=ftypes)
        fl = dlg.show()

        if fl != '':
            xsdentry.delete(0, END)
            xsdentry.insert(0, fl)


    def openepc():
        fl = tkFileDialog.asksaveasfilename(defaultextension=".epc")

        if fl != '':
            epcentry.delete(0, END)
            epcentry.insert(0, fl)


    def convert():
        status.set("Status: Converting")
        root.update()
        xmlpath = xmlentry.get()
        xsdpath = xsdentry.get()
        epcpath = epcentry.get()
        try:
            _convert(xmlpath, xsdpath, epcpath)
            status.set("Status: Success")
            tkMessageBox.showinfo("Converted", "Conversion finished!")
        except Exception as e:
            status.set("Status: Failed")
            tkMessageBox.showerror("Error", "An error has occurred: " + e.__str__())

    Button(root, text="Browse", command=openxml).grid(row=0, column=2)
    Button(root, text="Browse", command=openxsd).grid(row=1, column=2)
    Button(root, text="Browse", command=openepc).grid(row=2, column=2)

    Button(root, text="Convert", command=convert).grid(row=3, column=1)

    root.geometry("420x120+300+300")
    root.mainloop()

if __name__ == '__main__': main()
